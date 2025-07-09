from flask import Flask, request, render_template, redirect
from openpyxl import Workbook, load_workbook
from datetime import datetime
from dateutil.relativedelta import relativedelta
import os, datetime as dt


app = Flask(__name__)

def get_last_n_months(n):
    today = dt.date.today().replace(day=1)
    months = []

    for i in range(n):
        date = today - relativedelta(months=i)
        months.append(date.strftime('%b_%Y'))

    return months


@app.route('/')
def home():
    months = get_last_n_months(n=12)
    return render_template('home.html', months=months)


@app.route('/attendance', methods=['GET', 'POST'])
def get_details():
    results = []
    # get details to write to excel
    try:
        if request.method == 'POST':
            name = request.form.get('name')
            flat = request.form.get('flat')
            id = request.form.get('id')
            purpose = request.form.get('purpose')
            action = request.form.get('action')

            if not all([name, flat, id, purpose, action]):
                return "Missing form data. Please fill all fields.", 400
            
            now = dt.datetime.now()
            date_str = now.strftime('%Y-%m-%d')
            time_str = now.strftime('%H:%M:%S')
            
            if action == 'in':
                row = [name, flat, id, purpose, date_str, time_str, ""]
                write_to_sheet(row)
                return render_template('templates/success.html', amenity=f'{purpose} - Checked In')      
            elif action == 'out':
                updated = update_out_time(name, flat, id, purpose, date_str, time_str)
                if updated:
                    return render_template('templates/success.html', amenity=f'{purpose} - Checked Out')
                else:
                    return "No matching check-in found for today."
    except Exception as e:
        import traceback
        print("ERROR:", e)
        traceback.print_exc()
        return "Internal Server Error: " + str(e), 500

            
    # get details to search the excel
    query = request.args.get('query')
    amenity = request.args.get('amenity')
    month = request.args.get('month')

    if query and amenity and month:
        results = search_in_sheet(query, amenity, month)

    return render_template('templates/home.html', results=results)

def write_to_sheet(row_data):
    today = dt.date.today().strftime('%b_%Y')
    # file_map = {
    #     'gym': f'sheets/Gym_Attendance_{today}.xlsx',
    #     'carroms': f'sheets/Carroms_Attendance_{today}.xlsx',
    #     'pool': f'sheets/Pool_Attendance_{today}.xlsx',
    #     'tt': f'sheets/TT_Attendance_{today}.xlsx',
    # }

    file = f'sheets/Attendance_{today}.xlsx'

    if not os.path.exists(file):
        wb = Workbook()
        ws = wb.active
        ws.append(["NAME", "FLAT", "ID", "PURPOSE", "DATE", "IN-TIME", "OUT-TIME"])
        wb.save(file)

    
    wb = load_workbook(file)
    ws = wb.active
    ws.append(row_data)
    wb.save(file)


def update_out_time(name, flat, id, purpose, date_str, out_time):
    today = dt.date.today().strftime('%b_%Y')
    # file_map = {
    #     'gym': f'sheets/Gym_Attendance_{today}.xlsx',
    #     'carroms': f'sheets/Carroms_Attendance_{today}.xlsx',
    #     'pool': f'sheets/Pool_Attendance_{today}.xlsx',
    #     'tt': f'sheets/TT_Attendance_{today}.xlsx',
    # }

    file = f'sheets/Attendance_{today}.xlsx'
    
    wb = load_workbook(file)
    ws = wb.active
    updated = False
    for row in ws.iter_rows(min_row=2):  # skip header
        row_name = row[0].value
        row_flat = row[1].value
        row_id = row[2].value
        row_purpose = row[3].value
        row_date = row[4].value
        row_out = row[6].value

        if (row_name == name and row_flat == flat and row_id == id and row_purpose == purpose and row_date == date_str and not row_out):
            row[6].value = out_time
            updated = True
            break
    if updated:
        wb.save(file)
    return updated


def search_in_sheet(query, amenity, month):
    # file_map = {
    #     'gym': f'sheets/Gym_Attendance_{today}.xlsx',
    #     'carroms': f'sheets/Carroms_Attendance_{today}.xlsx',
    #     'pool': f'sheets/Pool_Attendance_{today}.xlsx',
    #     'tt': f'sheets/TT_Attendance_{today}.xlsx',
    # }
    file = f'sheets/Attendance_{month}.xlsx'

    if not os.path.exists(file):
        return []
    
    wb = load_workbook(file)
    ws = wb.active
    results = []


    for row in ws.iter_rows(min_row=2, values_only=True):
        if ((query.lower() in str(row[0]).lower() or query.lower() in str(row[1]).lower()) and amenity.lower() in str(row[3]).lower()):
            results.append(row)

    return results                 

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=8080)